"""
Parser for LinkedIn DEMOGRAPHICS sheet data.
Handles extraction and validation of demographic information about audience.
"""

from typing import Dict, List, Any, Optional
from dataclasses import dataclass


@dataclass
class DemographicItem:
    """Represents a single demographic entry with value and percentage"""
    name: str
    percentage: float

    def to_dict(self):
        return {
            "name": self.name,
            "percentage": self.percentage
        }


@dataclass
class DemographicsData:
    """Complete demographic information from the DEMOGRAPHICS sheet"""
    job_titles: List[DemographicItem]
    locations: List[DemographicItem]
    industries: List[DemographicItem]
    seniority: Optional[List[DemographicItem]] = None
    company_size: Optional[List[DemographicItem]] = None
    companies: Optional[List[DemographicItem]] = None

    def to_dict(self):
        return {
            "job_titles": [item.to_dict() for item in self.job_titles],
            "locations": [item.to_dict() for item in self.locations],
            "industries": [item.to_dict() for item in self.industries],
            "seniority": [item.to_dict() for item in self.seniority] if self.seniority else [],
            "company_size": [item.to_dict() for item in self.company_size] if self.company_size else [],
            "companies": [item.to_dict() for item in self.companies] if self.companies else [],
        }


def parse_percentage(value: Any) -> float:
    """
    Parse a percentage value from Excel cell (could be float, string, or percentage)

    Args:
        value: Value from Excel cell

    Returns:
        Parsed percentage as a float (0-1 or 0-100 depending on source)

    Raises:
        ValueError: If value cannot be converted to float
    """
    try:
        if isinstance(value, str):
            # Handle "< 1%" format
            if "< 1%" in value:
                return 0.005  # Approximate as 0.5%
            # Remove % symbol and spaces
            value = value.replace("%", "").replace("<", "").strip()

        percentage = float(value)
        return percentage
    except (ValueError, TypeError) as e:
        raise ValueError(f"Failed to parse percentage value '{value}': {str(e)}")


def parse_demographics_sheet(file_content: bytes) -> DemographicsData:
    """
    Parse the DEMOGRAPHICS sheet from LinkedIn Excel export.

    Expected structure:
    Row 1: Headers ["Top Demographics", "Value", "Percentage"]
    Row 2+: [Category, Item, Percentage]
            Categories: "Job titles", "Locations", "Industries", "Seniority", "Company size", "Companies"

    Args:
        file_content: Binary content of Excel file

    Returns:
        DemographicsData object with categorized demographic information

    Raises:
        ValueError: If DEMOGRAPHICS sheet not found or data is invalid
    """
    try:
        from openpyxl import load_workbook
        import io

        wb = load_workbook(io.BytesIO(file_content), data_only=True)

        if "DEMOGRAPHICS" not in wb.sheetnames:
            raise ValueError("DEMOGRAPHICS sheet not found in Excel file")

        ws = wb["DEMOGRAPHICS"]

        # Initialize storage for each demographic category
        demographics = {
            "Job titles": [],
            "Locations": [],
            "Industries": [],
            "Seniority": [],
            "Company size": [],
            "Companies": [],
        }

        # Skip header row (row 1)
        for row_idx in range(2, ws.max_row + 1):
            category_cell = ws.cell(row_idx, 1).value
            value_cell = ws.cell(row_idx, 2).value
            percentage_cell = ws.cell(row_idx, 3).value

            if category_cell and value_cell and percentage_cell is not None:
                category = str(category_cell).strip()
                value = str(value_cell).strip()

                try:
                    percentage = parse_percentage(percentage_cell)
                    item = DemographicItem(name=value, percentage=percentage)

                    if category in demographics:
                        demographics[category].append(item)

                except ValueError as e:
                    print(f"Warning: Could not parse demographic entry: {e}")
                    continue

        # Convert to DemographicsData object
        return DemographicsData(
            job_titles=demographics.get("Job titles", []),
            locations=demographics.get("Locations", []),
            industries=demographics.get("Industries", []),
            seniority=demographics.get("Seniority", []),
            company_size=demographics.get("Company size", []),
            companies=demographics.get("Companies", []),
        )

    except Exception as e:
        raise ValueError(f"Failed to parse DEMOGRAPHICS sheet: {str(e)}")
