"""
Parser for LinkedIn summary metrics calculations.
Calculates total engagements, average daily impressions, and new followers.
"""

from typing import Optional
from openpyxl import load_workbook
import io


def get_total_engagements(file_content: bytes) -> int:
    """
    Calculate total engagements by summing the 'Engagements' column in the ENGAGEMENT sheet.

    Args:
        file_content: Binary content of Excel file

    Returns:
        Total engagements as integer

    Raises:
        ValueError: If ENGAGEMENT sheet not found or Engagements column is missing
    """
    try:
        wb = load_workbook(io.BytesIO(file_content), data_only=True)

        if "ENGAGEMENT" not in wb.sheetnames:
            raise ValueError("ENGAGEMENT sheet not found in Excel file")

        ws = wb["ENGAGEMENT"]

        # Find the "Engagements" column header (usually in row 1)
        engagements_column = None
        for col_idx, cell in enumerate(ws[1], 1):
            if cell.value and str(cell.value).strip().lower() == "engagements":
                engagements_column = col_idx
                break

        if engagements_column is None:
            raise ValueError("'Engagements' column not found in ENGAGEMENT sheet")

        # Sum all values in the Engagements column (starting from row 2)
        total = 0
        for row_idx in range(2, ws.max_row + 1):
            cell = ws.cell(row_idx, engagements_column)
            if cell.value is not None:
                try:
                    total += float(cell.value)
                except (ValueError, TypeError):
                    # Skip cells that can't be converted to numbers
                    pass

        return int(total)

    except Exception as e:
        raise ValueError(f"Failed to calculate total engagements: {str(e)}")


def get_average_daily_impressions(total_impressions: int) -> float:
    """
    Calculate average daily impressions by dividing total impressions by 365.

    Args:
        total_impressions: Total impressions value

    Returns:
        Average daily impressions as float

    Raises:
        ValueError: If total_impressions is negative
    """
    if total_impressions < 0:
        raise ValueError("Total impressions cannot be negative")

    return total_impressions / 365.0


def get_new_followers(file_content: bytes) -> int:
    """
    Calculate total new followers by summing the 'New followers' column in the FOLLOWERS sheet.
    Starts from Row 4 (column headers are in row 3).

    Args:
        file_content: Binary content of Excel file

    Returns:
        Total new followers as integer

    Raises:
        ValueError: If FOLLOWERS sheet not found or 'New followers' column is missing
    """
    try:
        wb = load_workbook(io.BytesIO(file_content), data_only=True)

        if "FOLLOWERS" not in wb.sheetnames:
            raise ValueError("FOLLOWERS sheet not found in Excel file")

        ws = wb["FOLLOWERS"]

        # Find the "New followers" column header in row 3
        new_followers_column = None
        for col_idx, cell in enumerate(ws[3], 1):
            if cell.value and str(cell.value).strip().lower() == "new followers":
                new_followers_column = col_idx
                break

        if new_followers_column is None:
            raise ValueError("'New followers' column not found in Followers sheet")

        # Sum all values in the New followers column starting from row 4
        total = 0
        for row_idx in range(4, ws.max_row + 1):
            cell = ws.cell(row_idx, new_followers_column)
            if cell.value is not None:
                try:
                    total += float(cell.value)
                except (ValueError, TypeError):
                    # Skip cells that can't be converted to numbers
                    pass

        return int(total)

    except Exception as e:
        raise ValueError(f"Failed to calculate new followers: {str(e)}")


def calculate_summary_metrics(file_content: bytes, total_impressions: int) -> dict:
    """
    Calculate all summary metrics for the dashboard.

    Args:
        file_content: Binary content of Excel file
        total_impressions: Total impressions from discovery data

    Returns:
        Dictionary with calculated metrics:
        {
            "total_engagements": int,
            "average_impressions_per_day": float,
            "new_followers": int
        }

    Raises:
        ValueError: If any required sheet or column is missing
    """
    try:
        total_engagements = get_total_engagements(file_content)
        average_impressions_per_day = get_average_daily_impressions(total_impressions)
        new_followers = get_new_followers(file_content)

        return {
            "total_engagements": total_engagements,
            "average_impressions_per_day": average_impressions_per_day,
            "new_followers": new_followers
        }
    except ValueError as e:
        raise ValueError(f"Failed to calculate summary metrics: {str(e)}")
